import streamlit as st
import datetime
import pandas as pd
import openpyxl

book = openpyxl.Workbook()
#excelファイルの取得
book = openpyxl.load_workbook(filename='勤怠管理.xlsx')
#現在時刻の取得
time = datetime.datetime.today()
#シートの指定
sheetTime = book.worksheets[0]
#excel行
rowInt = 0
#excel列
columnInt = 0


st.title('出退勤管理アプリ')

if st.button("出勤"):
    columnInt = 1
    rowInt = 1
    while(not len(sheetTime.cell(row = rowInt, column = columnInt).value) is None):
        
    sheetTime.cell(row = rowInt, column = columnInt).value = time
    columnInt = 0
    rowInt = 0

if st.button("退勤"):
    columnInt += 4
    sheetTime.cell(row=rowInt,column=columnInt).value = time
    columnInt = 0

if st.button("休憩開始"):
    columnInt += 2
    sheetTime.cell(row=rowInt,column=columnInt).value = time
    columnInt = 0

if st.button("休憩終了"):
    columnInt += 3
    sheetTime.cell(row=rowInt,column=columnInt).value = time
    columnInt = 0

    
book.save('勤怠管理.xlsx')
book.close()





